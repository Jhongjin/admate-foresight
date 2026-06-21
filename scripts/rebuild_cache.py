"""
XLSX → JSON 캐시 재생성 스크립트
extractIndustry 로직을 Python으로 동일하게 구현하여 업종 재추출

노출위치 컬럼: XLSX에 publisher_platform + platform_position + impression_device 필요
  (Ads Manager → 보고서 → 세분화 → 게재 위치 추가 시 생성)
소재형태 컬럼: XLSX에 format 또는 소재 형태 컬럼 필요
  (Ads Manager → 보고서 → 세분화 → 소재 형태 추가 시 생성)
"""
import sys, json, re
sys.stdout.reconfigure(encoding='utf-8')
from pathlib import Path
import openpyxl

BASE = Path(__file__).parent.parent

# ── extractIndustry 로직 (xlsxLoader.ts와 동일) ──────────
VALID_INDUSTRIES = {
    '건설', '게임', '공공기관', '교육', '금융', '보험', '기관/단체',
    '방송통신', '병의원', '부동산', '뷰티', '생활/잡화', '서비스',
    '관광/레저', '수송', '식음료', '앱/사이트', '엔터테인먼트',
    '의약/건강식', '전자', '주류', '주택/가구', '패션', '기타',
}

INDUSTRY_NORMALIZE = {
    '기관.ver2': '기관/단체', '생활잡화': '생활/잡화', '생활잡화 공구': '생활/잡화',
    '식음료(에너지바)': '식음료', '의약': '의약/건기식', '의료/건강': '의약/건기식',
    '의료/건강(건강기능식품)': '의약/건기식', '의약품': '의약/건기식',
    '건강뷰티': '뷰티', '화장품': '뷰티', '기타기관': '기관/단체', '단체': '기관/단체',
    '기관': '기관/단체', '문화예술': '문화/예술', '앱서비스': '앱/사이트',
    '금융서비스': '금융', '건설/분양': '건설', '건설분양': '건설',
    '가전': '전자', '가전제품': '전자',
    '패션': '패션', '패션/의류': '패션', '패션/잡화': '패션', '의류': '패션',
    '의료기기': '병의원',
    '영화': '엔터테인먼트', '문화예술': '엔터테인먼트', '문화/예술': '엔터테인먼트',
    '에너지': '기타', '기타(에너지)': '기타',
    '공공기관 (신규)': '공공기관', '교육 (신규)': '교육', '말레이시아': '기타',
    '태국': '기타', '모마2': '기타', '모마': '기타', '광고쿠폰': '기타',
    '생화잡화': '생활/잡화', '생활잡화/안마의자': '생활/잡화', '주택가구': '주택/가구',
    '화장품(스킨케어)': '뷰티', '화장품/생활': '뷰티',
    '정부기관': '공공기관', '정부광고': '공공기관', '부동산/건설': '건설',
    '기타(국방업)': '기타', '쇼핑': '기타',
    '컴퓨터/기술': '전자', '박람회': '엔터테인먼트', '언론': '방송통신',
    '제조': '전자',
}

INDUSTRY_NORMALIZE.update({
    '의약': '의약/건강식', '의약/건기식': '의약/건강식',
    '의료/건강': '의약/건강식', '의료/건강(건강기능식품)': '의약/건강식',
    '의약품': '의약/건강식', '건강기능식품': '의약/건강식',
    '건강식품': '의약/건강식', '헬스케어': '의약/건강식',
    '여행': '관광/레저', '관광': '관광/레저', '레저': '관광/레저',
    '숙박': '관광/레저', '리조트': '관광/레저', '호텔': '관광/레저',
    '자동차': '수송', '운수': '수송', '항공': '수송', '자동차/수송': '수송',
    '부동산/건설': '부동산', '부동산/임대': '부동산', '임대': '부동산',
    '주류/음료': '주류', '맥주': '주류', '소주': '주류', '와인': '주류', '막걸리': '주류',
    'OTT': '방송통신', '방송': '방송통신', '통신': '방송통신', '미디어': '방송통신',
    '앱': '앱/사이트', '플랫폼': '앱/사이트', '커머스': '앱/사이트',
})

def normalize_industry(raw):
    value = (raw or '').strip()
    if not value:
        return '기타'
    mapped = INDUSTRY_NORMALIZE.get(value, value)
    return mapped if mapped in VALID_INDUSTRIES else '기타'

def is_valid_industry_part(s):
    if not s or len(s) > 20: return False
    if re.match(r'^\d+$', s): return False
    if s == 'Total': return False
    if '팀' in s or '본부' in s: return False
    if '협력광고' in s or 'Collaborative' in s: return False
    if '홍보' in s: return False
    if not re.search(r'[가-힣]', s): return False
    # 회사명 패턴: 한글(영문) 형태
    if re.search(r'[가-힣]+\([A-Za-z]', s): return False
    return True

def extract_industry(account_name):
    if not account_name: return '기타'
    parts = [p.strip() for p in account_name.strip().split('_') if p.strip()]
    for part in reversed(parts):
        if is_valid_industry_part(part):
            return normalize_industry(part)
    return '기타'

# ── 노출위치 매핑 ─────────────────────────────────────────
PLACEMENT_MAP = {
    ('facebook',        'feed'):              'FB 피드',
    ('facebook',        'right_hand_column'): 'FB 우측 컬럼',
    ('facebook',        'story'):             'FB 스토리',
    ('facebook',        'reels'):             'FB 릴스',
    ('facebook',        'video_feeds'):       'FB 동영상 피드',
    ('facebook',        'search'):            'FB 검색',
    ('facebook',        'marketplace'):       'FB 마켓플레이스',
    ('facebook',        'groups_feed'):       'FB 그룹 피드',
    ('instagram',       'stream'):            'IG 피드',
    ('instagram',       'story'):             'IG 스토리',
    ('instagram',       'reels'):             'IG 릴스',
    ('instagram',       'explore'):           'IG 탐색 탭',
    ('instagram',       'explore_home'):      'IG 탐색 홈',
    ('instagram',       'profile_feed'):      'IG 프로필 피드',
    ('audience_network','classic'):           'AN 네이티브',
    ('audience_network','rewarded_video'):    'AN 리워드 동영상',
    ('audience_network','instream_video'):    'AN 인스트림',
    ('messenger',       'messenger_home'):    'MSG 홈',
    ('messenger',       'story'):             'MSG 스토리',
    ('messenger',       'sponsored_messages'):'MSG 스폰서',
}

DEVICE_MAP = {
    'desktop':    'PC',
    'mobile_app': 'MO',
    'mobile_web': 'MO',
    'tablet':     'TB',
}

OPTIMIZATION_GOAL_MAP = {
    'NONE':                              '자동최적화',
    'Unknown Optimization Goal':         '알수없음',
    'IMPRESSIONS':                       '노출',
    'REACH':                             '도달',
    'LINK_CLICKS':                       '링크클릭',
    'LANDING_PAGE_VIEWS':                '랜딩페이지조회',
    'THRUPLAY':                          '스루플레이',
    'TWO_SECOND_CONTINUOUS_VIDEO_VIEWS': '2초동영상조회',
    'POST_ENGAGEMENT':                   '게시물참여',
    'OFFSITE_CONVERSIONS':               '전환',
    'RETURN_ON_AD_SPEND':                '광고수익률',
    'LEAD_GENERATION':                   '리드',
    'QUALITY_LEAD':                      '양질의리드',
    'APP_INSTALLS':                      '앱설치',
    'ENGAGED_USERS':                     '앱참여',
    'VALUE':                             '구매가치',
    'CONVERSATIONS':                     '대화',
    'REPLIES':                           '답장',
    'REMINDERS_SET':                     '리마인더설정',
    'VISIT_INSTAGRAM_PROFILE':           'IG프로필방문',
    'SOCIAL_IMPRESSIONS':                '소셜노출',
    'EVENT_RESPONSES':                   '이벤트응답',
    'PAGE_LIKES':                        '페이지좋아요',
    'AD_RECALL_LIFT':                    '광고회상',
}

FORMAT_MAP = {
    'image':      '이미지',
    'video':      '동영상',
    'carousel':   '슬라이드',
    'collection': '컬렉션',
    'IMAGE':      '이미지',
    'VIDEO':      '동영상',
    'CAROUSEL':   '슬라이드',
    'COLLECTION': '컬렉션',
}

def build_placement(platform, position, device):
    platform = (platform or '').strip().lower()
    position = (position or '').strip().lower()
    device   = (device   or '').strip().lower()
    base     = PLACEMENT_MAP.get((platform, position))
    if not base:
        base = f'{platform}/{position}' if platform else ''
    device_label = DEVICE_MAP.get(device, '')
    return f'{base} ({device_label})' if base and device_label else base

# ── cost_per_action_type 파싱 ─────────────────────────────
def parse_action_map(raw):
    result = {}
    if not raw: return result
    for item in str(raw).split(','):
        item = item.strip()
        if ':' in item:
            k, _, v = item.partition(':')
            try: result[k.strip()] = float(v.strip())
            except: pass
    return result

# ── XLSX 읽기 ─────────────────────────────────────────────
xlsx_path = BASE / 'data' / 'Meta_Total_Accounts_Daily_Report_0330_01.xlsx'
print(f'XLSX 읽는 중: {xlsx_path.name}')
wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
ws = wb.active

headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
h = {v: i for i, v in enumerate(headers) if v}
print(f'총 컬럼: {len(headers)}, 헤더 일부: {headers[:10]}')

def col(row, name, default=''):
    i = h.get(name)
    if i is None: return default
    v = row[i]
    return v if v is not None else default

def flt(v, default=0.0):
    try: return float(v)
    except: return default

records = []
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    reach = flt(col(row, 'reach'))
    impressions = flt(col(row, 'impressions'))
    spend = flt(col(row, 'spend'))
    cpm = flt(col(row, 'cpm'))

    if reach <= 0 or impressions <= 0 or cpm <= 0:
        continue

    cost_map = parse_action_map(col(row, 'cost_per_action_type'))

    노출위치 = build_placement(
        col(row, 'publisher_platform'),
        col(row, 'platform_position'),
        col(row, 'impression_device'),
    )

    소재형태_raw = str(col(row, 'format', col(row, '소재 형태', col(row, 'creative_type', '')))).strip()
    소재형태 = FORMAT_MAP.get(소재형태_raw, 소재형태_raw)

    records.append({
        '업종':        extract_industry(str(col(row, 'ad_account_name', ''))),
        '캠페인이름':  str(col(row, 'campaign_name', '')).strip(),
        '목표':        str(col(row, 'objective', '')).strip(),
        '최적화목표':  OPTIMIZATION_GOAL_MAP.get(str(col(row, 'optimization_goal', '')).strip(), str(col(row, 'optimization_goal', '')).strip()),
        '노출위치':    노출위치,
        '소재형태':    소재형태,
        '성별':        str(col(row, 'gender', '')).strip(),
        '연령':        str(col(row, 'age', '')).strip(),
        '도달':        reach,
        '노출':        impressions,
        '지출금액':    spend,
        '빈도':        flt(col(row, 'frequency')),
        'CPM':         cpm,
        'CPC':         flt(col(row, 'cpc')),
        'CPC링크':     cost_map.get('link_click', 0),
        '영상조회수':  flt(col(row, 'video_view', 0) if 'video_view' in h else 0),
        '영상조회비용':cost_map.get('video_view', 0),
        '날짜':        str(col(row, 'date_start', '')).strip(),
    })

    if i % 50000 == 0:
        print(f'  {i:,}행 처리 중...')

wb.close()
print(f'필터 후 총 {len(records):,}행')

# ── JSON 저장 ─────────────────────────────────────────────
out_path = BASE / 'data' / 'ad_data_cache.json'
out_path.write_text(json.dumps(records, ensure_ascii=False), encoding='utf-8')
print(f'JSON 캐시 저장 완료: {out_path}')

# ── 업종 분포 확인 ────────────────────────────────────────
from collections import Counter
ind_counter = Counter(r['업종'] for r in records)
print('\n업종 분포 (상위 30):')
for k, v in ind_counter.most_common(30):
    print(f'  {repr(k)}: {v:,}건')
